"""
excel_report.py
---------------
Exporta o resultado da regressão para uma planilha Excel com abas de dados,
ajustados, métricas, calibração e fórmulas plug-and-play.

Portado de `vdm-regressao-previsto/pipeline.py`.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

import numpy as np
import pandas as pd


def export_to_excel(result: dict) -> bytes:
    """Serializa o resultado do pipeline em um workbook Excel (bytes)."""
    df_full = result["df_result"].copy()
    metrics_df = result["metrics_df"]
    clip_log = result["clip_log"]
    targets = list(result.get("fonte_cols", {}).keys())

    ajust_cols = [f"{t}_ajustado" for t in targets if f"{t}_ajustado" in df_full.columns]

    fonte_to_drop = [c for c in df_full.columns if c.startswith("fonte_")]
    df_full.drop(columns=fonte_to_drop, inplace=True)

    if "fonte" in df_full.columns:
        df_full["fonte"] = df_full["fonte"].apply(
            lambda v: "observado" if v == "observado" else "previsto"
        )

    bool_pattern = re.compile(r".+_is_.+")
    bool_cols = [c for c in df_full.columns if bool_pattern.fullmatch(c) and c not in ajust_cols]
    df_full.drop(columns=bool_cols, inplace=True)

    df_dados = df_full.drop(columns=[c for c in ajust_cols if c in df_full.columns])
    df_obs = df_dados[df_dados["fonte"] == "observado"].copy()

    df_ajust = pd.DataFrame()
    if ajust_cols:
        has_fitted = df_full[ajust_cols].notna().any(axis=1)
        df_base = df_full[has_fitted].copy()
        base_cols = [c for c in df_base.columns if c not in targets and c not in ajust_cols]
        df_ajust = df_base[base_cols].copy()

        for t in targets:
            adj_col = f"{t}_ajustado"
            obs_col = f"{t}_observado"
            if t in df_base.columns:
                df_ajust[obs_col] = df_base[t].values
            if adj_col in df_base.columns:
                df_ajust[adj_col] = df_base[adj_col].values
                if obs_col in df_ajust.columns:
                    obs_s = df_ajust[obs_col]
                    adj_s = df_ajust[adj_col]
                    df_ajust[f"{t}_residuo"] = (obs_s - adj_s).round(4)
                    df_ajust[f"{t}_erro_pct"] = (
                        ((obs_s - adj_s) / obs_s.replace(0, np.nan) * 100).round(2)
                    )

        group_col = result.get("group_col", "regional_macro")
        if group_col in df_ajust.columns:
            for t in targets:
                obs_col = f"{t}_observado"
                adj_col = f"{t}_ajustado"
                if obs_col not in df_ajust.columns or adj_col not in df_ajust.columns:
                    continue
                conf_col = f"{t}_confianca"
                df_ajust[conf_col] = "alta"
                for reg in df_ajust[group_col].dropna().unique():
                    reg_mask = df_ajust[group_col].astype(str) == str(reg)
                    obs_reg = df_ajust.loc[reg_mask, obs_col].dropna()
                    if len(obs_reg) < 5:
                        continue
                    p10 = np.percentile(obs_reg, 10)
                    p90 = np.percentile(obs_reg, 90)
                    adj_reg = df_ajust.loc[reg_mask, adj_col]
                    outside = (adj_reg < p10) | (adj_reg > p90)
                    df_ajust.loc[reg_mask & outside, conf_col] = "baixa"

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_dados.to_excel(writer, sheet_name="dados", index=False)
        df_obs.to_excel(writer, sheet_name="observados", index=False)
        if not df_ajust.empty:
            df_ajust.to_excel(writer, sheet_name="ajustados_obs", index=False)
        if not metrics_df.empty:
            metrics_df.to_excel(writer, sheet_name="metricas_modelos", index=False)
        if clip_log:
            pd.DataFrame(clip_log).to_excel(writer, sheet_name="diagnostico_zeros", index=False)
        if not df_ajust.empty:
            _build_calibracao_sheet(writer, targets, df_ajust, metrics_df, result)
            _apply_conditional_formatting(writer, targets)
        _build_equacoes_sheet(writer, targets, result, df_dados)

    return buf.getvalue()


def _build_calibracao_sheet(writer, targets, df_ajust, metrics_df, result):
    rows = []
    for t in targets:
        obs_col = f"{t}_observado"
        adj_col = f"{t}_ajustado"
        if obs_col not in df_ajust.columns or adj_col not in df_ajust.columns:
            continue
        obs = df_ajust[obs_col].dropna()
        adj = df_ajust[adj_col].dropna()
        common = obs.index.intersection(adj.index)
        obs = obs.loc[common]
        adj = adj.loc[common]
        n = len(obs)
        if n == 0:
            continue
        err_pct = ((obs - adj) / obs.replace(0, np.nan) * 100).abs()
        metrics_rl = metrics_df[
            (metrics_df["metodo"] == "regressao_linear") & (metrics_df["target"] == t)
        ] if not metrics_df.empty else pd.DataFrame()
        r2_mean = float(metrics_rl["r2"].dropna().mean()) if (not metrics_rl.empty and not metrics_rl["r2"].dropna().empty) else None
        rows.append({
            "Target": t,
            "N observacoes": n,
            "R² medio (regioes)": round(r2_mean, 4) if r2_mean else None,
            "RMSE": round(float(np.sqrt(np.mean((obs - adj) ** 2))), 2),
            "MAE": round(float(np.mean(np.abs(obs - adj))), 2),
            "MAPE (%)": round(float(err_pct.mean()), 1),
            "Mediana |erro%|": round(float(err_pct.median()), 1),
            "% |erro| <= 35%": round(100 * (err_pct <= 35).sum() / n, 1),
            "% |erro| <= 50%": round(100 * (err_pct <= 50).sum() / n, 1),
            "% |erro| <= 80%": round(100 * (err_pct <= 80).sum() / n, 1),
            "Pior |erro%|": round(float(err_pct.max()), 1),
            "Predicoes zeradas": int((adj == 0).sum()),
        })
    pd.DataFrame(rows).to_excel(writer, sheet_name="calibracao", index=False)


def _build_equacoes_sheet(writer, targets, result, df_dados):
    equations_df = result.get("equations_df", pd.DataFrame())
    if equations_df.empty:
        pd.DataFrame({"info": ["Sem equacoes disponiveis"]}).to_excel(
            writer, sheet_name="equacoes", index=False
        )
        return

    col_map = {col: i for i, col in enumerate(df_dados.columns)}

    def col_letter(idx):
        s = ""
        while idx >= 0:
            s = chr(65 + (idx % 26)) + s
            idx = idx // 26 - 1
        return s

    KNOWN_CATEGORICAL = ["classe", "situacao", "perim_urb", "trecho"]

    def resolve_feature_ref(feat):
        for cat in KNOWN_CATEGORICAL:
            if feat.startswith(cat + "_"):
                value = feat[len(cat) + 1:]
                if cat in col_map:
                    cl = col_letter(col_map[cat])
                    return f'IF(dados!{cl}2="{value}",1,0)'
        if feat in col_map:
            cl = col_letter(col_map[feat])
            return f"dados!{cl}2"
        return f"[{feat}]"

    rows = []
    group_col = result.get("group_col", "regional_macro")
    for t in targets:
        eq_rows = equations_df[equations_df["target"] == t]
        if eq_rows.empty:
            continue
        for _, eq_row in eq_rows.iterrows():
            reg = str(eq_row.get(group_col, ""))
            eq_text = eq_row["equation"]
            formula_parts = []
            feature_refs = []
            for line in eq_text.split("\n"):
                line = line.strip()
                if not line or "Duan smearing" in line or "exp(log" in line:
                    continue
                if " = " in line and not formula_parts:
                    rhs = line.split(" = ", 1)[1]
                    m = re.match(r"([+-]?\s*[\d.]+)", rhs)
                    if m:
                        val = float(m.group(1).replace(" ", ""))
                        formula_parts.append(f"{val:.4f}".replace(".", ","))
                    continue
                m = re.match(r"([+-])\s+([\d.]+)\s+×\s+(.+)", line)
                if m:
                    sign, coef, feat = m.group(1), float(m.group(2)), m.group(3).strip()
                    ref = resolve_feature_ref(feat)
                    formula_parts.append(f"{sign} {abs(coef):.4f}*{ref}".replace(".", ","))
                    feature_refs.append(f"{feat} → {ref}")
            if formula_parts:
                formula = "=" + " ".join(formula_parts)
                label = f"{t}" if reg == "global" else f"{t} (Reg {reg})"
                nota = ""
                if result.get("log_transform"):
                    sm = eq_row.get("smearing")
                    nota = (f"Resultado em log. Aplique: ŷ = EXP(formula) * {sm} "
                            f"(Duan smearing)") if sm else "Resultado em log: ŷ = EXP(formula)"
                rows.append({
                    "Target/Regiao": label,
                    "Formula Excel (linha 2 — arraste para baixo)": formula,
                    "Pos-processamento": nota,
                    "Features referenciadas": "\n".join(feature_refs) if feature_refs else "",
                    "Equacao original": eq_text,
                })
    pd.DataFrame(rows).to_excel(writer, sheet_name="equacoes", index=False)


def _apply_conditional_formatting(writer, targets):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    ws = writer.sheets["ajustados_obs"]
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cell in ws[1]:
        if cell.value and "_erro_pct" in str(cell.value):
            cl = cell.column_letter
            rng = f"{cl}2:{cl}1048576"
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["-35", "35"], fill=green))
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["-80", "-35.01"], fill=yellow))
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["35.01", "80"], fill=yellow))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-80.01"], fill=red))
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["80"], fill=red))


def export_filename() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"sre_previsao_trafego_{ts}.xlsx"
